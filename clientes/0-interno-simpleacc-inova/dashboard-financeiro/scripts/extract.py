#!/usr/bin/env python3
"""Extrai a planilha 'DFC e DRE Financeiro Simple Acc' para o dados.json
que o dashboard consome.

Uso local:
    python3 scripts/extract.py caminho/para/planilha.xlsm

Sem argumento, procura 'dfc_dre.xlsm' ao lado do script.
A sincronizacao diaria (GitHub Actions) chama extrair() via scripts/sync.py.
"""
import sys, os, json, datetime, re

MESES = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
         'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']


def _parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return datetime.date(v.year, v.month, v.day)
    if isinstance(v, str):
        m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{4})', v.strip())
        if m:
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return datetime.date(y, mo, d)
            except ValueError:
                return None
    return None


def extrair(xlsx_path, gerado_em=None):
    """Le a planilha e devolve o dict que vira dados.json."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- Lancamentos (fonte da verdade) ---
    ws = wb['Lancamentos']
    lanc = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        data, tipo, cat, sub, desc, parte, status, valor = r[0:8]
        if not tipo or not isinstance(valor, (int, float)):
            continue
        d = _parse_date(data)
        lanc.append({
            'data': d.isoformat() if d else None,
            'mes': d.month if d else None,
            'tipo': str(tipo).strip(),
            'categoria': str(cat).strip() if cat else 'Outros',
            'sub': str(sub).strip() if sub else '',
            'descricao': str(desc).strip() if desc else '',
            'parte': str(parte).strip() if parte else '',
            'status': str(status).strip() if status else '',
            'valor': round(float(valor), 2),
        })

    # --- DRE Mensal oficial da planilha (referencia) ---
    dre = {}
    if 'DRE_Mensal' in wb.sheetnames:
        wd = wb['DRE_Mensal']
        drows = list(wd.iter_rows(values_only=True))
        hdr = drows[2] if len(drows) > 2 else []
        meses_cols = [(i, h) for i, h in enumerate(hdr) if h in MESES]
        for row in drows[3:]:
            linha = row[0]
            if not linha:
                continue
            dre[str(linha)] = {h: (row[i] or 0) for i, h in meses_cols}

    return {
        'gerado_em': gerado_em or datetime.date.today().isoformat(),
        'fonte': 'Planilha DFC/DRE Financeiro Simple Acc',
        'meses': MESES,
        'lancamentos': lanc,
        'dre_mensal': dre,
    }


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else \
        os.path.join(os.path.dirname(__file__), 'dfc_dre.xlsm')
    out = extrair(path)
    dest = os.path.join(os.path.dirname(__file__), '..', 'dados.json')
    with open(dest, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    print('OK: %d lancamentos -> %s' % (len(out['lancamentos']), os.path.abspath(dest)))


if __name__ == '__main__':
    main()
